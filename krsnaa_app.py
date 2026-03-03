import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from num2words import num2words
import io
import zipfile
import sqlite3
from datetime import datetime
import calendar

# --- CONFIG ---
DB_NAME = "krsnaa_billing_v2.db"
COMPANY_NAME = "KRSNAA RETAIL PVT LTD"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS invoices 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  bill_no TEXT, 
                  bill_date TEXT, 
                  subcentre TEXT, 
                  category TEXT, 
                  amount REAL)''')
    conn.commit()
    conn.close()

def get_last_bill_no():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT bill_no FROM invoices ORDER BY id DESC LIMIT 1")
    result = c.fetchone()
    conn.close()
    if result:
        try:
            return int(result[0].split('/')[-1])
        except:
            return 0
    return 0

# --- LOGIN ---
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if not st.session_state["authenticated"]:
        st.markdown(f"<h2 style='text-align: center; color: #1E3A8A;'>{COMPANY_NAME}</h2>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.write("### 🔐 Secure Login")
            user = st.text_input("Username")
            pw = st.text_input("Password", type="password")
            if st.button("Login"):
                if user == "admin" and pw == "krsnaa@2026":
                    st.session_state["authenticated"] = True
                    st.rerun()
                else:
                    st.error("❌ Invalid Credentials")
        return False
    return True

init_db()

if check_password():
    st.set_page_config(page_title=COMPANY_NAME, layout="wide")

    try:
        st.sidebar.image("logo.png", width=200) 
    except:
        st.sidebar.subheader(COMPANY_NAME)
    
    st.sidebar.markdown("---")
    page = st.sidebar.radio("Go To", ["🏠 Dashboard", "🗄️ Billing Records"])

    if page == "🏠 Dashboard":
        st.markdown(f"""
            <style>
            .centered-title {{ text-align: center; color: #1E3A8A; font-size: 42px; font-weight: bold; padding-bottom: 10px; text-shadow: 1px 1px 2px #d1d1d1; }}
            .subtitle {{ text-align: center; font-size: 18px; color: #555; margin-bottom: 30px; }}
            </style>
            <h1 class="centered-title">{COMPANY_NAME}</h1>
            <p class="subtitle">Automated Invoice Generation System</p>
            """, unsafe_allow_html=True)

        st.markdown("---")
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.subheader("⚙️ Settings")
            inv_date = st.date_input("Select Invoice Date", datetime.now())
            
            formatted_date = inv_date.strftime('%d-%m-%Y')
            month_name = inv_date.strftime('%B') 
            year_val = inv_date.strftime('%Y')   
            period_text = f"{month_name}-{year_val}" 
            narration_text = f"Narration - Invoice for the period of {month_name} {year_val}" 

            last_num = get_last_bill_no()
            start_no = st.number_input("Starting Bill No", min_value=1, value=last_num + 1)
            p_file = st.file_uploader("Upload Data (Excel)", type=["xlsx"])
            t_file = st.file_uploader("Upload Template (Excel)", type=["xlsx"])

        with col2:
            if p_file:
                df = pd.read_excel(p_file) 
                st.success(f"File Uploaded! Total {len(df)} records found.")
                st.dataframe(df.head(10), use_container_width=True)
                
                if st.button("🚀 Generate & Save All Invoices"):
                    if t_file:
                        grouped = list(df.groupby(['SubCentreName', 'Sub Category'])) # List mein convert kiya length nikalne ke liye
                        total_invoices = len(grouped)
                        
                        # --- PROGRESS BAR START ---
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        # --------------------------

                        zip_buffer = io.BytesIO()
                        current_bill = start_no
                        conn = sqlite3.connect(DB_NAME)
                        
                        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                            for i, ((sub, cat), group) in enumerate(grouped):
                                # Progress update (0.0 to 1.0)
                                percent_complete = (i + 1) / total_invoices
                                progress_bar.progress(percent_complete)
                                status_text.text(f"Processing: {i+1}/{total_invoices} - {sub}")
                                bill_str = f"KDL/INV/25-26/{current_bill:03d}"
                                amt_col = 'Hospital Credit' if 'Hospital Credit' in group.columns else 'Amount'
                                net_amt = int(group[amt_col].sum())
                                inv_to_col = 'Invoice To' if 'Invoice To' in group.columns else 'Invoice to'
                                
                                t_file.seek(0)
                                wb = load_workbook(t_file)
                                ws = wb.active
                                
                                ws["A10"] = group[inv_to_col].iloc[0]
                                ws["A11"] = cat
                                ws["A12"] = sub
                                ws["A13"] = group['Address'].iloc[0]
                                ws["E10"] = bill_str
                                ws["C20"] = len(group)
                                ws["E20"] = net_amt
                                ws["A44"] = f"Amount (in Words): {num2words(net_amt, lang='en_IN').title()} Rupees Only"
                                ws["E9"] = formatted_date
                                ws["E11"] = period_text      
                                ws["A45"] = narration_text   
                                
                                if "Deatial" in wb.sheetnames:
                                    d_ws = wb["Deatial"]
                                    if d_ws.max_row > 1: d_ws.delete_rows(2, d_ws.max_row)
                                    headers = [d_ws.cell(row=1, column=c).value for c in range(1, 25) if d_ws.cell(row=1, column=c).value]
                                    last_row = 2
                                    for r_idx, (_, row) in enumerate(group.iterrows(), start=2):
                                        for c_idx, h in enumerate(headers, start=1):
                                            if h in group.columns:
                                                d_ws.cell(row=r_idx, column=c_idx).value = row[h]
                                        last_row = r_idx
                                    d_ws.cell(row=last_row + 2, column=1).value = "Note: Certified that all above patients are shown on the NVHCP Portal"
                                
                                f_stream = io.BytesIO()
                                wb.save(f_stream)
                                zip_file.writestr(f"Invoice_{current_bill:03d}.xlsx", f_stream.getvalue())
                                
                                conn.execute("INSERT INTO invoices (bill_no, bill_date, subcentre, category, amount) VALUES (?,?,?,?,?)",
                                             (bill_str, formatted_date, sub, cat, net_amt))
                                current_bill += 1
                        
                        conn.commit()
                        conn.close()
                        # Finish hone par progress bar hata dena ya success message dena
                        progress_bar.empty() 
                        status_text.success(f"✅ Sabhi {current_bill - start_no} Invoices taiyar hain!") # Success message
                        st.success(f"✅ Generated {current_bill - start_no} Invoices.")
                        st.download_button("📥 Download ZIP Package", zip_buffer.getvalue(), f"KRSNAA_Bills_{formatted_date}.zip")
                    else:
                        st.error("Please upload the template file!")

    elif page == "🗄️ Billing Records":
        st.header("Billing History")
        conn = sqlite3.connect(DB_NAME)
        records_df = pd.read_sql_query("SELECT * FROM invoices ORDER BY id DESC", conn)
        conn.close()
        
        if not records_df.empty:
            st.write("Select records or delete all:")
            
            # Action Buttons Layout
            act_col1, act_col2 = st.columns([1, 1])
            
            with act_col1:
                selected_ids = st.multiselect("Select IDs to Delete", records_df['id'].tolist())
                if st.button("🗑️ Delete Selected"):
                    if selected_ids:
                        conn = sqlite3.connect(DB_NAME)
                        query = f"DELETE FROM invoices WHERE id IN ({','.join(map(str, selected_ids))})"
                        conn.execute(query)
                        conn.commit()
                        conn.close()
                        st.success(f"Deleted {len(selected_ids)} records.")
                        st.rerun()
                    else:
                        st.warning("Please select at least one ID.")

            with act_col2:
                st.write("⚠️ **Danger Zone**")
                confirm_delete_all = st.checkbox("Confirm: I want to delete EVERYTHING")
                if st.button("🔥 DELETE ALL RECORDS"):
                    if confirm_delete_all:
                        conn = sqlite3.connect(DB_NAME)
                        conn.execute("DELETE FROM invoices")
                        conn.execute("DELETE FROM sqlite_sequence WHERE name='invoices'") # Reset ID counter
                        conn.commit()
                        conn.close()
                        st.error("All billing records have been wiped out!")
                        st.rerun()
                    else:
                        st.info("Check the confirmation box first.")

            st.markdown("---")
            st.dataframe(records_df, use_container_width=True)
        else:
            st.info("Database is empty.")

    if st.sidebar.button("Logout"):
        st.session_state["authenticated"] = False
        st.rerun()